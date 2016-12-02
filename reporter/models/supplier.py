from __future__ import unicode_literals
import json
from django.db import models
from django.db.models.signals import pre_save, post_save
from django.dispatch import receiver
from reporter.models import Ecu 
from django.utils.dateparse import parse_datetime
from openpyxl import load_workbook


class TrwFailsafeMatrix(models.Model):
    file = models.FileField(upload_to='static/uploads/supplier/trw/failsafe/')
    ecu = models.ForeignKey('reporter.Ecu')
    class Meta:
        verbose_name = 'TRW Failsafe Matrix'
        verbose_name_plural = 'TRW Failsafe Matrices'
    def __unicode__(self):
        return unicode(self.file)
    
    @receiver(pre_save, sender='reporter.TrwFailsafeMatrix')
    def parse_file(sender, instance, signal, **kwargs):
        matrix_file = load_workbook(filename=instance.file, read_only=True)
        # Create DTC
        dtc_worksheet = matrix_file['DET - Work view']
        for row in dtc_worksheet.iter_rows(row_offset=1):
            if row[5].value:
                for fc in row[3].value.split():
                    dtc, created = Dtc.objects.get_or_create(
                        code=row[5].value[row[5].value.index('0x')+2:],
                        ecu=instance.ecu,
                        name=fc
                    )
